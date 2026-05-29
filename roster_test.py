import json
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook


AFFINITY_PATH = Path(r"C:\Users\User\Downloads\affinity_index.json")
WORKBOOK_PATH = Path(r"C:\Users\User\Downloads\삼국지14_전체무장_공유(통합본)의 사본.xlsx")

REL_VALUE = {"☆": 3, "◎": 2, "○": 1}
REL_LABEL = {3: "☆", 2: "◎", 1: "○"}
POSITIVE = set(REL_VALUE)
TRAIT_COLOR_SCORE = {
    "FFFFF2CC": 5,  # gold
    "FFC9DAF8": 2,  # blue/general
    "FFF4CCCC": 0,  # red
}
EXTRA_TRAITS = {"악주", "응원"}


def display_name(officer, ambiguous_names):
    name = officer["name"]
    return f"{name}#{officer['id']}" if name in ambiguous_names else name


def load_affinity():
    with AFFINITY_PATH.open("r", encoding="utf-8") as f:
        data = json.load(f)

    officers = {}
    by_name = defaultdict(list)
    for raw in data["characters"]:
        officer = {
            "id": int(raw["id"]),
            "name": raw["name"],
            "support": bool(raw.get("support_flag")),
        }
        officers[officer["id"]] = officer
        by_name[officer["name"]].append(officer["id"])

    directed = {}
    for edge in data["edges"]:
        rel = edge["relation"]
        if rel not in POSITIVE:
            continue
        src = int(edge["source_id"])
        for tid in edge.get("target_ids", []):
            directed[(src, int(tid))] = rel

    return officers, by_name, directed


def load_sheet_data():
    wb = load_workbook(WORKBOOK_PATH, data_only=False)
    ws = wb["전체무장 능력치개성"]
    raw = wb["반에디터_원본"]

    stats = {}
    for row in range(3, ws.max_row + 1):
        officer_id = ws.cell(row, 1).value
        if officer_id is None:
            continue
        officer_id = int(officer_id)
        values = {
            "통솔": ws.cell(row, 3).value or 0,
            "무력": ws.cell(row, 4).value or 0,
            "지력": ws.cell(row, 5).value or 0,
            "정치": ws.cell(row, 6).value or 0,
            "매력": ws.cell(row, 7).value or 0,
        }
        trait_score = 0
        traits = []
        for col in range(9, 14):
            cell = ws.cell(row, col)
            trait = cell.value
            if not trait:
                continue
            color = cell.fill.fgColor.rgb
            base = TRAIT_COLOR_SCORE.get(color, 0)
            extra = 1 if trait in EXTRA_TRAITS else 0
            trait_score += base + extra
            traits.append(str(trait))
        values["trait_score"] = trait_score
        values["traits"] = traits
        values["primary"] = max(values["통솔"] + values["무력"], values["통솔"] + values["지력"]) + trait_score
        values["secondary"] = values["통솔"] + values["정치"] + trait_score
        values["third"] = values["지력"] + values["정치"] + trait_score
        values["fourth"] = values["통솔"] + trait_score
        stats[officer_id] = values

    # Raw sheet columns known from inspection: A/B/C include id/name/gender.
    gender = {}
    spouse = {}
    for row in range(2, raw.max_row + 1):
        officer_id = raw.cell(row, 1).value
        if officer_id is None:
            continue
        try:
            officer_id = int(officer_id)
        except (TypeError, ValueError):
            continue
        gender[officer_id] = raw.cell(row, 3).value
        spouse[officer_id] = raw.cell(row, 30).value

    return stats, gender, spouse


def pair_relation(a, b, directed):
    ab = directed.get((a, b))
    ba = directed.get((b, a))
    best = max([REL_VALUE.get(ab, 0), REL_VALUE.get(ba, 0)])
    if not best:
        return None
    if ab and ba:
        arrow = "<->"
    elif ab:
        arrow = "->"
    else:
        arrow = "<-"
    return arrow, REL_LABEL[best], best


def group_metrics(group, directed):
    buff = {oid: 0 for oid in group}
    edges = []
    ids = list(group)
    for i, a in enumerate(ids):
        for b in ids[i + 1 :]:
            rel = pair_relation(a, b, directed)
            if not rel:
                continue
            arrow, label, value = rel
            buff[a] += 1
            buff[b] += 1
            edges.append((a, b, arrow, label, value))
    return buff, edges


def best_oath_triples(group, directed, stats):
    triples = []
    ids = list(group)
    for i, a in enumerate(ids):
        for j in range(i + 1, len(ids)):
            b = ids[j]
            for c in ids[j + 1 :]:
                pairs = [(a, b), (a, c), (b, c)]
                existing = sum(1 for x, y in pairs if pair_relation(x, y, directed))
                if existing == 3:
                    continue
                # If oath makes a full triangle, every missing pair adds +1 to both endpoints.
                gain = (3 - existing) * 2
                power = sum(stats.get(x, {}).get("primary", 0) for x in (a, b, c))
                triples.append((existing, gain, power, (a, b, c)))
    triples.sort(key=lambda x: (x[0], x[1], x[2]), reverse=True)
    return triples[:5]


def group_score(group, directed, stats):
    buff, edges = group_metrics(group, directed)
    power = sum(stats.get(oid, {}).get("primary", 0) for oid in group)
    stack = sum(buff.values())
    edge_quality = sum(e[4] for e in edges)
    return power + stack * 15 + edge_quality * 10, buff, edges


def pick_roster(input_names, officers, by_name, directed, stats):
    fixed = {}
    ambiguous = {}
    for name in input_names:
        candidates = by_name.get(name, [])
        if len(candidates) == 1:
            fixed[name] = candidates[0]
        elif len(candidates) > 1:
            ambiguous[name] = candidates

    selected = dict(fixed)
    for name, candidates in ambiguous.items():
        best = None
        for cid in candidates:
            degree_to_pool = 0
            for other_id in fixed.values():
                if pair_relation(cid, other_id, directed):
                    degree_to_pool += 1
            primary = stats.get(cid, {}).get("primary", 0)
            score = degree_to_pool * 1000 + primary
            if best is None or score > best[0]:
                best = (score, cid)
        selected[name] = best[1]
    return selected, ambiguous


def print_group(title, group, officers, ambiguous_names, directed, stats):
    score, buff, edges = group_score(group, directed, stats)
    print(f"\n[{title}] score={score:.1f}, members={len(group)}")
    ranked = sorted(group, key=lambda oid: (buff[oid], stats.get(oid, {}).get("primary", 0)), reverse=True)
    print("  구성:", ", ".join(
        f"{display_name(officers[oid], ambiguous_names)}(P{stats.get(oid, {}).get('primary', 0)},B{buff[oid]})"
        for oid in ranked
    ))
    if edges:
        print("  연결:")
        for a, b, arrow, label, _ in sorted(edges, key=lambda e: (-e[4], display_name(officers[e[0]], ambiguous_names), display_name(officers[e[1]], ambiguous_names))):
            print(f"    {display_name(officers[a], ambiguous_names)} {arrow}{label} {display_name(officers[b], ambiguous_names)}")
    else:
        print("  연결: 없음")
    triples = best_oath_triples(group, directed, stats)
    if triples:
        print("  의형제 후보:")
        for existing, gain, power, triple in triples[:3]:
            names = " + ".join(display_name(officers[x], ambiguous_names) for x in triple)
            print(f"    {names} / 기존연결 {existing}/3, 예상 스택 +{gain}, P합 {power}")
    return score


def main():
    input_names = "순욱, 곽가, 유엽, 태사자, 공융, 양수, 황충, 법정, 유비, 관우, 장비, 공손찬, 후씨, 조운".replace(" ", "").split(",")
    officers, by_name, directed = load_affinity()
    stats, gender, spouse = load_sheet_data()
    selected_by_name, ambiguous = pick_roster(input_names, officers, by_name, directed, stats)
    selected = selected_by_name
    ambiguous_names = {name for name in input_names if len(by_name.get(name, [])) > 1}

    print("선택 ID:")
    for name in input_names:
        oid = selected[name]
        stat = stats.get(oid, {})
        sex = gender.get(oid, "?")
        traits = ",".join(stat.get("traits", []))
        print(f"  {name} -> {display_name(officers[oid], ambiguous_names)} / 성별={sex} / P={stat.get('primary', 0)} / 개성점수={stat.get('trait_score', 0)} / {traits}")

    if ambiguous:
        print("\n동명이인:")
        for name, ids in ambiguous.items():
            for oid in ids:
                stat = stats.get(oid, {})
                linked = []
                for other_name, other_id in selected.items():
                    if other_name == name:
                        continue
                    rel = pair_relation(oid, other_id, directed)
                    if rel:
                        linked.append(f"{display_name(officers[other_id], ambiguous_names)}{rel[0]}{rel[1]}")
                print(f"  {name} 후보 {oid}: P={stat.get('primary', 0)}, 개성={','.join(stat.get('traits', []))}, 명단연결={', '.join(linked) or '없음'}")

    ids = {name: selected[name] for name in input_names}
    scenarios = [
        (
            "A. 촉 핵심을 한 분파로 묶음",
            [
                ["유비", "관우", "장비", "조운", "황충", "법정", "공손찬", "후씨"],
                ["공융", "태사자", "양수"],
                ["순욱", "곽가", "유엽"],
            ],
        ),
        (
            "B. 관우 독립 시도",
            [
                ["관우"],
                ["유비", "장비", "조운", "황충", "법정", "공손찬", "후씨"],
                ["공융", "태사자", "양수"],
                ["순욱", "곽가", "유엽"],
            ],
        ),
        (
            "C. 장비-황충-법정 분리 시도",
            [
                ["유비", "관우", "조운", "공손찬", "후씨"],
                ["장비", "황충", "법정"],
                ["공융", "태사자", "양수"],
                ["순욱", "곽가", "유엽"],
            ],
        ),
    ]

    print("\n시나리오 비교")
    for scenario_title, groups in scenarios:
        print(f"\n## {scenario_title}")
        total = 0
        for idx, names in enumerate(groups, 1):
            group_ids = [ids[name] for name in names]
            total += print_group(f"분파 {idx}", group_ids, officers, ambiguous_names, directed, stats)
        print(f"  => total={total:.1f}")


if __name__ == "__main__":
    main()

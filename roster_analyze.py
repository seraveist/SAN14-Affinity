import json
from collections import defaultdict, deque
from itertools import combinations
from pathlib import Path

from openpyxl import load_workbook


AFFINITY_PATH = Path(r"C:\Users\User\Downloads\affinity_index.json")
WORKBOOK_PATH = Path(r"C:\Users\User\Downloads\삼국지14_전체무장_공유(통합본)의 사본.xlsx")

REL_VALUE = {"☆": 3, "◎": 2, "○": 1}
REL_LABEL = {3: "☆", 2: "◎", 1: "○"}
POSITIVE = set(REL_VALUE)
TRAIT_COLOR_SCORE = {"FFFFF2CC": 5, "FFC9DAF8": 2, "FFF4CCCC": 0}
EXTRA_TRAITS = {"악주", "응원"}

ROSTER_TEXT = (
    "유비,관우,장비,황충,조운,마초,마대,제갈량,서서,방통,마속,주창,관평,관흥,관은병,"
    "엄안,법정,석도,공손찬,후씨,감씨,미씨,황월영,마운록,노식,주준,황보숭,양씨,아귀,"
    "등지,진도,요화,위연,이엄,이회"
)


def load_affinity():
    with AFFINITY_PATH.open("r", encoding="utf-8") as f:
        data = json.load(f)
    officers = {}
    by_name = defaultdict(list)
    for raw in data["characters"]:
        officer = {
            "id": int(raw["id"]),
            "name": raw["name"],
            "support": bool(raw.get("support_flag")),
        }
        officers[officer["id"]] = officer
        by_name[officer["name"]].append(officer["id"])

    directed = {}
    for edge in data["edges"]:
        rel = edge["relation"]
        if rel not in POSITIVE:
            continue
        src = int(edge["source_id"])
        for tid in edge.get("target_ids", []):
            directed[(src, int(tid))] = rel
    return officers, by_name, directed


def load_sheet_data():
    wb = load_workbook(WORKBOOK_PATH, data_only=False)
    ws = wb["전체무장 능력치개성"]
    raw = wb["반에디터_원본"]

    stats = {}
    for row in range(3, ws.max_row + 1):
        officer_id = ws.cell(row, 1).value
        if officer_id is None:
            continue
        officer_id = int(officer_id)
        values = {
            "통솔": ws.cell(row, 3).value or 0,
            "무력": ws.cell(row, 4).value or 0,
            "지력": ws.cell(row, 5).value or 0,
            "정치": ws.cell(row, 6).value or 0,
            "매력": ws.cell(row, 7).value or 0,
        }
        traits = []
        trait_score = 0
        for col in range(9, 14):
            cell = ws.cell(row, col)
            trait = cell.value
            if not trait:
                continue
            color = cell.fill.fgColor.rgb
            score = TRAIT_COLOR_SCORE.get(color, 0)
            if trait in EXTRA_TRAITS:
                score += 1
            trait_score += score
            traits.append(str(trait))
        values["traits"] = traits
        values["trait_score"] = trait_score
        values["primary"] = max(values["통솔"] + values["무력"], values["통솔"] + values["지력"]) + trait_score
        values["secondary"] = values["통솔"] + values["정치"] + trait_score
        values["third"] = values["지력"] + values["정치"] + trait_score
        values["fourth"] = values["통솔"] + trait_score
        stats[officer_id] = values

    gender = {}
    spouse = {}
    for row in range(2, raw.max_row + 1):
        officer_id = raw.cell(row, 1).value
        if officer_id is None:
            continue
        try:
            officer_id = int(officer_id)
        except (TypeError, ValueError):
            continue
        gender[officer_id] = raw.cell(row, 3).value
        spouse[officer_id] = raw.cell(row, 30).value
    return stats, gender, spouse


def pair_relation(a, b, directed):
    ab = directed.get((a, b))
    ba = directed.get((b, a))
    best = max(REL_VALUE.get(ab, 0), REL_VALUE.get(ba, 0))
    if not best:
        return None
    if ab and ba:
        arrow = "<->"
    elif ab:
        arrow = "->"
    else:
        arrow = "<-"
    return arrow, REL_LABEL[best], best


def display(officers, oid, ambiguous_names):
    name = officers[oid]["name"]
    return f"{name}#{oid}" if name in ambiguous_names else name


def resolve_roster(names, officers, by_name, directed, stats):
    fixed = {}
    ambiguous = {}
    for name in names:
        ids = by_name.get(name, [])
        if len(ids) == 1:
            fixed[name] = ids[0]
        elif len(ids) > 1:
            ambiguous[name] = ids
        else:
            fixed[name] = None

    selected = dict(fixed)
    fixed_ids = [oid for oid in fixed.values() if oid is not None]
    for name, ids in ambiguous.items():
        best = None
        for oid in ids:
            linked = sum(1 for other in fixed_ids if pair_relation(oid, other, directed))
            score = linked * 1000 + stats.get(oid, {}).get("primary", 0)
            if best is None or score > best[0]:
                best = (score, oid)
        selected[name] = best[1]
    return selected, ambiguous


def group_metrics(group, directed, stats):
    buff = {oid: 0 for oid in group}
    edges = []
    for a, b in combinations(group, 2):
        rel = pair_relation(a, b, directed)
        if rel:
            arrow, label, value = rel
            buff[a] += 1
            buff[b] += 1
            edges.append((a, b, arrow, label, value))
    power = sum(stats.get(oid, {}).get("primary", 0) for oid in group)
    stack = sum(buff.values())
    quality = sum(edge[4] for edge in edges)
    score = power + stack * 15 + quality * 10
    return score, buff, edges


def top_oaths(group, directed, stats, limit=5):
    rows = []
    for a, b, c in combinations(group, 3):
        pairs = [(a, b), (a, c), (b, c)]
        existing = sum(1 for x, y in pairs if pair_relation(x, y, directed))
        if existing == 3:
            continue
        # Prefer triangles with two existing links. One-link triangles are shown only if very useful.
        gain = (3 - existing) * 2
        power = sum(stats.get(x, {}).get("primary", 0) for x in (a, b, c))
        rows.append((existing, gain, power, (a, b, c)))
    rows.sort(key=lambda row: (row[0], row[2], row[1]), reverse=True)
    return rows[:limit]


def top_marriages(group, directed, stats, gender, limit=5):
    # Female officers are treated as married in the base formation.
    # This section is a what-if view: if female officers were unmarried, which new links help most?
    rows = []
    for a, b in combinations(group, 2):
        ga, gb = gender.get(a), gender.get(b)
        if {ga, gb} != {"남", "여"}:
            continue
        male = a if ga == "남" else b
        female = b if ga == "남" else a
        rel = pair_relation(male, female, directed)
        existing = 1 if rel else 0
        female_bonus = 2 if any(t in EXTRA_TRAITS for t in stats.get(female, {}).get("traits", [])) else 0
        gain = (0 if existing else 2) + female_bonus
        power = stats.get(male, {}).get("primary", 0) + stats.get(female, {}).get("primary", 0)
        rows.append((existing, gain, power, male, female))
    rows.sort(key=lambda row: (row[0] == 0, row[1], row[2]), reverse=True)
    return rows[:limit]


def print_group(title, group, officers, ambiguous_names, directed, stats, gender):
    score, buff, edges = group_metrics(group, directed, stats)
    print(f"\n[{title}] score={score:.1f}, members={len(group)}")
    ranked = sorted(group, key=lambda oid: (buff[oid], stats.get(oid, {}).get("primary", 0)), reverse=True)
    print("  구성:", ", ".join(
        f"{display(officers, oid, ambiguous_names)}(P{stats.get(oid, {}).get('primary', 0)},B{buff[oid]})"
        for oid in ranked
    ))
    print("  연결:")
    for a, b, arrow, label, value in sorted(edges, key=lambda e: (-e[4], display(officers, e[0], ambiguous_names), display(officers, e[1], ambiguous_names))):
        print(f"    {display(officers, a, ambiguous_names)} {arrow}{label} {display(officers, b, ambiguous_names)}")
    if not edges:
        print("    없음")
    print("  의형제 후보:")
    for existing, gain, power, triple in top_oaths(group, directed, stats, 3):
        names = " + ".join(display(officers, oid, ambiguous_names) for oid in triple)
        print(f"    {names} / 기존연결 {existing}/3, 예상스택 +{gain}, P합 {power}")
    print("  결혼 후보(미혼 가정):")
    for existing, gain, power, male, female in top_marriages(group, directed, stats, gender, 3):
        ex = "기존친애 있음" if existing else "신규연결"
        print(f"    {display(officers, male, ambiguous_names)} + {display(officers, female, ambiguous_names)} / {ex}, 예상가치 +{gain}, P합 {power}")
    return score


def main():
    names = [name.strip() for name in ROSTER_TEXT.split(",") if name.strip()]
    officers, by_name, directed = load_affinity()
    stats, gender, spouse = load_sheet_data()
    selected_by_name, ambiguous = resolve_roster(names, officers, by_name, directed, stats)
    ambiguous_names = {name for name in names if len(by_name.get(name, [])) > 1}

    print("선택 ID:")
    for name in names:
        oid = selected_by_name.get(name)
        if oid is None:
            print(f"  {name}: NOT FOUND")
            continue
        stat = stats.get(oid, {})
        print(
            f"  {name} -> {display(officers, oid, ambiguous_names)} / "
            f"성별={gender.get(oid)} / P={stat.get('primary')} / 개성점수={stat.get('trait_score')} / "
            f"{','.join(stat.get('traits', []))}"
        )

    if ambiguous:
        print("\n동명이인:")
        for name, ids in ambiguous.items():
            for oid in ids:
                linked = []
                for other_name, other_id in selected_by_name.items():
                    if other_id is None or other_name == name:
                        continue
                    rel = pair_relation(oid, other_id, directed)
                    if rel:
                        linked.append(f"{display(officers, other_id, ambiguous_names)}{rel[0]}{rel[1]}")
                stat = stats.get(oid, {})
                print(f"  {name} 후보 {oid}: P={stat.get('primary')}, 개성={','.join(stat.get('traits', []))}, 명단연결={', '.join(linked) or '없음'}")

    ids = selected_by_name
    groups = [
        ("1분파 유비-제갈량 군사/내정", ["유비", "제갈량", "서서", "방통", "마속", "황월영", "감씨", "미씨", "석도", "이엄"]),
        ("2분파 관우-장비-조운/관씨 직계", ["관우", "장비", "조운", "관평", "관흥", "관은병", "주창", "요화", "진도", "등지"]),
        ("3분파 황충-법정/마초 서량", ["황충", "엄안", "법정", "마초", "마대", "마운록", "양씨", "아귀", "이회", "위연"]),
        ("4분파 공손찬-노식 한말 라인", ["공손찬", "후씨", "노식", "주준", "황보숭"]),
    ]
    print("\n분파 샘플")
    total = 0
    for title, group_names in groups:
        group = [ids[name] for name in group_names if ids.get(name) is not None]
        total += print_group(title, group, officers, ambiguous_names, directed, stats, gender)
    print(f"\n총점={total:.1f}")


if __name__ == "__main__":
    main()

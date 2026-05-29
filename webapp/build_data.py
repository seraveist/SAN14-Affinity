import json
import os
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent
REPO_ROOT = ROOT.parent
LEGACY_AFFINITY_PATH = Path(r"C:\Users\User\Downloads\affinity_index.json")
LEGACY_WORKBOOK_PATH = Path(r"C:\Users\User\Downloads\삼국지14_전체무장_공유(통합본)의 사본.xlsx")
OUT = ROOT / "data.js"

TRAIT_COLOR = {
    "FFFFF2CC": "gold",
    "FFC9DAF8": "blue",
    "FFF4CCCC": "red",
}
TRAIT_SCORE = {
    "gold": 5,
    "blue": 2,
    "red": 0,
}
EXTRA_TRAITS = {"악주", "응원"}
POSITIVE = {"☆", "◎", "○"}


def resolve_input_path(env_name, repo_filename, legacy_path):
    candidates = []
    env_value = os.environ.get(env_name)
    if env_value:
        candidates.append(Path(env_value))
    candidates.extend([REPO_ROOT / repo_filename, legacy_path])

    for path in candidates:
        if path.exists():
            return path

    checked = "\n".join(f"- {path}" for path in candidates)
    raise FileNotFoundError(f"{env_name} input file not found. Checked:\n{checked}")


AFFINITY_PATH = resolve_input_path(
    "SAN14_AFFINITY_JSON",
    "affinity_index.json",
    LEGACY_AFFINITY_PATH,
)
WORKBOOK_PATH = resolve_input_path(
    "SAN14_WORKBOOK",
    "삼국지14_전체무장_공유(통합본)의 사본.xlsx",
    LEGACY_WORKBOOK_PATH,
)


def load_affinity():
    with AFFINITY_PATH.open("r", encoding="utf-8") as f:
        data = json.load(f)

    officers = {}
    by_name = defaultdict(list)
    raw_by_id = {}
    for raw in data["characters"]:
        oid = int(raw["id"])
        raw_by_id[oid] = raw
        officers[oid] = {
            "id": oid,
            "name": raw["name"],
            "support": bool(raw.get("support_flag")),
        }
        by_name[raw["name"]].append(oid)

    def disambiguate_targets(edge):
        target_ids = [int(target) for target in edge.get("target_ids", [])]
        if len(target_ids) <= 1:
            return target_ids

        source_name = edge.get("source_name")
        matched = []
        for target_id in target_ids:
            relations = raw_by_id.get(target_id, {}).get("relations", {})
            names = [name for rows in relations.values() for name in rows]
            if source_name in names:
                matched.append(target_id)

        return matched or target_ids

    edges = []
    for edge in data["edges"]:
        relation = edge["relation"]
        if relation not in POSITIVE:
            continue
        source = int(edge["source_id"])
        for target in disambiguate_targets(edge):
            edges.append({
                "source": source,
                "target": int(target),
                "relation": relation,
            })
    return officers, by_name, edges


def load_workbook_data(officers):
    wb = load_workbook(WORKBOOK_PATH, data_only=False)
    ws = wb["전체무장 능력치개성"]
    raw = wb["반에디터_원본"]

    for row in range(3, ws.max_row + 1):
        oid = ws.cell(row, 1).value
        if oid is None:
            continue
        oid = int(oid)
        officer = officers.get(oid)
        if officer is None:
            continue

        stats = {
            "leadership": ws.cell(row, 3).value or 0,
            "war": ws.cell(row, 4).value or 0,
            "intelligence": ws.cell(row, 5).value or 0,
            "politics": ws.cell(row, 6).value or 0,
            "charm": ws.cell(row, 7).value or 0,
        }

        traits = []
        trait_score = 0
        for col in range(9, 14):
            cell = ws.cell(row, col)
            if not cell.value:
                continue
            color = TRAIT_COLOR.get(cell.fill.fgColor.rgb, "none")
            score = TRAIT_SCORE.get(color, 0)
            if cell.value in EXTRA_TRAITS:
                score += 1
            trait_score += score
            traits.append({
                "name": str(cell.value),
                "color": color,
                "score": score,
            })

        officer["stats"] = stats
        officer["traits"] = traits
        officer["traitScore"] = trait_score
        officer["primaryScore"] = max(
            stats["leadership"] + stats["war"],
            stats["leadership"] + stats["intelligence"],
        ) + trait_score
        officer["secondaryScore"] = stats["leadership"] + stats["politics"] + trait_score
        officer["thirdScore"] = stats["intelligence"] + stats["politics"] + trait_score
        officer["fourthScore"] = stats["leadership"] + trait_score

    for row in range(2, raw.max_row + 1):
        oid = raw.cell(row, 1).value
        if oid is None:
            continue
        try:
            oid = int(oid)
        except (TypeError, ValueError):
            continue
        officer = officers.get(oid)
        if officer is None:
            continue
        officer["gender"] = raw.cell(row, 3).value
        officer["spouse"] = raw.cell(row, 30).value


def main():
    officers, by_name, edges = load_affinity()
    load_workbook_data(officers)

    payload = {
        "metadata": {
            "title": "삼국지14 장수 편성기",
            "relationTypes": ["☆", "◎", "○"],
            "femaleDefaultMarried": True,
        },
        "officers": officers,
        "byName": dict(sorted(by_name.items())),
        "edges": edges,
    }

    text = json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
    with OUT.open("w", encoding="utf-8", newline="\n") as out:
        out.write(f"window.SAN14_DATA = {text};\n")
    print(f"wrote {OUT} ({OUT.stat().st_size:,} bytes)")


if __name__ == "__main__":
    main()
